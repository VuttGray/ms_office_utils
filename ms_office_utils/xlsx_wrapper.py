from json import dump
from logging import getLogger
from re import sub

from openpyxl import cell as oxl_cell
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = getLogger('logger')


def get_cell_value(cell):
    if cell.value:
        return ("'" if cell.style_array.quotePrefix == 1 else "") + str(cell.value).strip()
    return ""


class XlsxManager:
    def __init__(self, file_path):
        self.__wb = load_workbook(filename=file_path)
        self.__ws = self.__wb.active

    def close(self):
        self.__wb.close()

    def __open_sheet(self, sheet_name):
        self.__ws = self.__wb[sheet_name]

    def __find_column(self, header):
        for i in range(1, 100):
            if self.__ws.cell(1, i).value == header:
                return i

    def __find_empty_row(self, col_index, start_row_index=2):
        return self.__find_value(col_index, "", start_row_index)

    def __find_value(self, col_index, value, start_row_index=2):
        for i in range(start_row_index, 3500):
            val = self.__ws.cell(i, col_index).value
            if val == value:
                return i
            if value == "" and val is None:
                return i
        raise UserWarning("Unexpected break: too many rows")

    def save(self, filename):
        self.__wb.save(filename)

    def update_value(self, row_index, col_index, value):
        self.__ws.cell(self, row_index, col_index).value = value

    def update_values(self, row_index, values):
        for column in values.keys():
            col_index = self.__find_column(column)
            if col_index == -1:
                raise UserWarning("Unknown column: " + column)
            self.__ws.cell(row_index, col_index).value = values[column]

    def update_datarow(self, key_column, key_value, values):
        key_col_index = self.__find_column(key_column)
        key_row_index = self.__find_value(key_col_index, key_value)
        if key_row_index == -1:
            key_row_index = self.__find_empty_row(key_col_index)
            self.update_value(key_row_index, key_col_index, key_value)
        self.update_values(key_row_index, values)


class XlsxReadOnlyManager:
    def __init__(self, file_path):
        try:
            self.__wb = load_workbook(filename=file_path, read_only=True)
            self.__ws = self.__wb.active
        except InvalidFileException as ex:
            logger.error(f'Excel loading failed: {file_path}')
            logger.exception(ex)
            return
        except Exception as ex:
            logger.error(f'Excel loading failed: {file_path}')
            logger.exception(ex)
            return

    def find_sheet_name(self, name_to_search):
        for sh in self.__wb.sheetnames:
            if sh.lower() == name_to_search.lower():
                return sh

    def open_sheet(self, sheet_name):
        self.__ws = self.__wb[sheet_name]

    def close(self):
        self.__wb.close()

    def __find_empty_row(self, col_index, start_row_index=2):
        return self.__find_value(col_index, "", start_row_index)

    def __find_value(self, col_index, value, start_row_index=2):
        for i, row in enumerate(self.__ws.rows):
            if i < start_row_index:
                continue
            for cell in row:
                if isinstance(cell, oxl_cell.read_only.EmptyCell):
                    if value == "":
                        return i
                    else:
                        continue
                if col_index not in [cell.column, cell.column_letter]:
                    break
                if cell.value == value:
                    return i
                if value == "" and cell.value is None:
                    return i
        return -1

    def get_headers(self, columns, row_index, prepare_value=False):
        if row_index < 0:
            raise ValueError
        cell_values = {}
        for row in self.get_rows(row_index, row_index):
            for cell in row:
                if isinstance(cell, oxl_cell.read_only.EmptyCell):
                    if columns is None:
                        break
                    else:
                        continue
                if columns is None or cell.column_letter in columns:
                    value = cell.value.strip()
                    if prepare_value:
                        value = value.replace('  ', ' ').replace(' ', '_')
                        value = sub(r'[\W]', '', value)
                    cell_values[cell.column_letter] = value
        return cell_values

    def get_headers_list(self, columns, row_index):
        cell_values = self.get_headers(columns, row_index, True)
        if cell_values:
            return ', '.join(cell_values.values())
        else:
            return ', '.join(['col{0}'.format(i) for i, ci in enumerate(columns)])

    def generate_sql_select_query(self, sheet_name, column_letters, first_row_index, last_row_index=-1):
        self.open_sheet(sheet_name)
        rows = []
        for row_i, row in enumerate(self.__ws.rows):
            if row_i < first_row_index:
                continue
            if row_i > last_row_index > 0:
                break
            cell_values = []
            for cell in row:
                if isinstance(cell, oxl_cell.read_only.EmptyCell):
                    continue
                if cell.column_letter in column_letters:
                    cell_values.append(cell.value)
            if not cell_values or all(val is None for val in cell_values):
                continue
            rows.append(",".join(["'{0}'".format(val) if val else "null" for val in cell_values]))
        values = "(" + ")\n\t\t, (".join(rows) + ")"
        headers = self.get_headers_list(column_letters, first_row_index - 1)
        return 'select {0} \nfrom (values {1}\n\t) t ({0})'.format(headers, values)

    def get_rows(self, first_row_index, last_row_index):
        if self.__ws:
            for row_i, row in enumerate(self.__ws.rows):
                if row_i < first_row_index:
                    continue
                if row_i > last_row_index:
                    break
                yield row

    def get_data(self, first_row_index, last_row_index, first_column_index, last_column_index, f_headers, t_headers):
        data = list(dict())
        for row in self.get_rows(first_row_index, last_row_index):
            fixed_entry = dict()
            entry = dict()
            for col_i, cell in enumerate(row):
                if col_i < first_column_index:
                    continue
                if col_i > last_column_index:
                    break
                if isinstance(cell, oxl_cell.read_only.EmptyCell):
                    continue
                if cell.column_letter in f_headers.keys():
                    fixed_entry[f_headers[cell.column_letter]] = str(cell.value).strip()
                    continue
                if cell.column_letter in t_headers.keys():
                    title_column = t_headers[cell.column_letter]['title_column']
                    title = t_headers[cell.column_letter]['title']
                    value_column = t_headers[cell.column_letter]['value_column']
                    if not entry:
                        entry = dict(fixed_entry)
                    elif entry[title_column] != title:
                        data.append(entry)
                        entry = dict(fixed_entry)
                    entry[title_column] = title
                    entry[value_column] = str(cell.value).strip()
            if entry:
                data.append(entry)
            else:
                data.append(fixed_entry)
        return data

    def get_cells(self, first_row_index, last_row_index, first_column_index, last_column_index):
        for row in self.get_rows(first_row_index, last_row_index):
            for col_i, cell in enumerate(row):
                if col_i < first_column_index:
                    continue
                if col_i > last_column_index:
                    break
                yield cell


def read_data(file_path: str,
              fixed_columns_number: int,
              transposed_column_name: str,
              default_value_column: str):
    xl = XlsxReadOnlyManager(file_path)
    headers = xl.get_headers(None, 0)
    fixed_headers = {}
    transposed_headers = {}
    for i, (k, v) in enumerate(headers.items()):
        if i < fixed_columns_number or fixed_columns_number == -1:
            fixed_headers[k] = v
        else:
            header_parts = v.split(' - ')
            if len(header_parts) == 1:
                value_column = default_value_column
                title = v
            else:
                value_column = header_parts[-1].strip()
                title = ' - '.join(header_parts[:-1]).strip()
            transposed_headers[k] = {'title_column': transposed_column_name,
                                     'title': title,
                                     'value_column': value_column}
    return xl.get_data(1, 100000, 0, len(headers), fixed_headers, transposed_headers)


def excel2json(xl_path: str,
               json_path: str,
               fixed_columns_number: int = -1,
               transposed_column_name: str = "Repeated column",
               default_value_column: str = "Value") -> None:
    """
    Read data from a table in Excel file and export the result as a json file
    :param xl_path: path to the source Excel file
    :param json_path: path to the target json file
    :param fixed_columns_number: number of fixed columns, the columns that will be exported as main attributes
    :param transposed_column_name: the union name for repeated (transposed) columns
    :param default_value_column: the default name for value attribute of repeated (transposed) columns
    other columns will be transposed and exported as different rows.
    Specify -1 if you what export all columns as a main attribute (header)
    :return: None

    example: excel2json(r'C:\SomeFolder\Subjects.xlsx', r'C:\SomeFolder\subjects.json', 1, 'Visit', 'Visit date')
    """
    data = read_data(xl_path, fixed_columns_number, transposed_column_name, default_value_column)
    with open(json_path, 'w', encoding="utf-8") as fp:
        dump(data, fp, indent=4, ensure_ascii=False)
